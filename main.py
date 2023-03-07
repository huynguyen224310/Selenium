import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from pathlib import Path
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook
import os
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

base_path = Path(__file__).parent
s = Service(f'{base_path}\chromedriver.exe')
browser = webdriver.Chrome(service=s)
book = openpyxl.load_workbook(f'{base_path}\Order.xlsx')
sheet = book.active
def login():
    id_ex = sheet.cell(row=28, column=1).value
    pw_ex = sheet.cell(row=29, column=1).value
    browser.get('https://khachhang.giaohangtietkiem.vn/khach-hang/dang_nhap')
    username = browser.find_element(By.XPATH, '/html/body/div/header/nav/div/div/div[2]/div/div/div/form/div[2]/div[1]/div[1]/input')
    username.send_keys(id_ex)
    pw = browser.find_element(By.XPATH, '/html/body/div/header/nav/div/div/div[2]/div/div/div/form/div[2]/div[2]/div[1]/input')
    pw.send_keys(pw_ex)
    dangnhap = browser.find_element(By.XPATH,'/html/body/div/header/nav/div/div/div[2]/div/div/div/form/div[2]/div[3]/button/span')
    dangnhap.click()
    time.sleep(2.5)
    browser.get('https://khachhang.giaohangtietkiem.vn/web/tao-don-hang/don-excel')
    time.sleep(2)
    thietlap = browser.find_element(By.XPATH,'/html/body/div[1]/main/div[2]/div[2]/div/div[2]/button')
    thietlap.click()
    time.sleep(0.75)
    # shop_tra_ship = browser.find_element(By.XPATH,'/html/body/div[3]/div/div[2]/div/div[2]/div[2]/div[1]/div[4]/div/label[1]/span[1]')
    # shop_tra_ship.click()
    # time.sleep(0.75)
    # goi_shop = browser.find_element(By.XPATH,'/html/body/div[3]/div/div[2]/div/div[2]/div[2]/div[1]/div[6]/div/div[6]/div[1]/label/span[2]')
    # goi_shop.click()
    # time.sleep(0.75)
    ca_lay = browser.find_element(By.XPATH,'/html/body/div[2]/div/div[2]/div/div[2]/div[2]/div[1]/div[1]/div[2]/div')
    ca_lay.click()
    time.sleep(0.5)
    ca_toi = browser.find_element(By.XPATH,'/html/body/div[3]/div/div/div/div/div/div/div/label[1]')
    ca_toi.click()
    xac_nhan = browser.find_element(By.XPATH,'/html/body/div[2]/div/div[2]/div/div[2]/div[3]/div/button[2]')
    xac_nhan.click()
    time.sleep(2)

def order():
    for i in range(1,11):
        Name_EX = sheet.cell(row = i , column=1).value
        Name = browser.find_element(By.XPATH,f'/html/body/div[1]/main/div[2]/div[2]/div/div[3]/div/div[2]/div[{i}]/div/div/div[4]/div/div/input')
        if Name_EX == 'ok':
            break
        else:
            Name.send_keys(str(Name_EX))
        SDT_EX = sheet.cell(row=i,column=2).value
        SDT = browser.find_element(By.XPATH,f'/html/body/div[1]/main/div[2]/div[2]/div/div[3]/div/div[2]/div[{i}]/div/div/div[5]/div/input')
        SDT.send_keys('0' + str(SDT_EX))
        # Dia chi
        DC_EX = sheet.cell(row = i,column=3).value
        DC = browser.find_element(By.XPATH,f'/html/body/div[1]/main/div[2]/div[2]/div/div[3]/div/div[2]/div[{i}]/div/div/div[6]/div/div[1]/input')
        DC.send_keys(str(DC_EX))
        time.sleep(0.5)
        DC.click()
        time.sleep(0.5)
        cli = browser.find_element(By.XPATH,
                                   '/html/body/div[1]/main/div[2]/div[1]/span[1]')
        time.sleep(0.5)
        cli.click()
        # SP
        SP_EX = sheet.cell(row=i, column=4).value
        SP = browser.find_element(By.XPATH,
                                  f'/html/body/div[1]/main/div[2]/div[2]/div/div[3]/div/div[2]/div[{i}]/div/div/div[7]/div/div[2]/div[1]/div/input')
        SP.send_keys(str(SP_EX))
        # So luong
        SL_EX = sheet.cell(row=i, column=5).value
        SL = browser.find_element(By.XPATH,
                                  f'/html/body/div[1]/main/div[2]/div[2]/div/div[3]/div/div[2]/div[{i}]/div/div/div[7]/div/div[2]/div[2]/input')
        SL.send_keys(Keys.BACKSPACE + str(SL_EX))
        # Khoi luong
        khoiluong_ex = sheet.cell(row=i, column=6).value
        KL = browser.find_element(By.XPATH,
                                  f'/html/body/div[1]/main/div[2]/div[2]/div/div[3]/div/div[2]/div[{i}]/div/div/div[7]/div/div[2]/div[3]/input')
        KL.send_keys(Keys.BACKSPACE * 4 + str(khoiluong_ex))
        # Gia tri hang
        value_ex = sheet.cell(row=i, column=7).value
        Gia_tri = browser.find_element(By.XPATH,
                                       f'/html/body/div[1]/main/div[2]/div[2]/div/div[3]/div/div[2]/div[{i}]/div/div/div[7]/div/div[2]/div[4]/input')
        Gia_tri.send_keys(Keys.BACKSPACE + str(value_ex) + '000')
        # COD
        COD_EX = sheet.cell(row=i, column=8).value
        COD = browser.find_element(By.XPATH,
                                   f'/html/body/div[1]/main/div[2]/div[2]/div/div[3]/div/div[2]/div[{i}]/div/div/div[8]/div/input')
        COD.send_keys(Keys.BACKSPACE + str(COD_EX) + '000')


if __name__ == '__main__':
        login()
        order()